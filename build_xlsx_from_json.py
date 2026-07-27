#!/usr/bin/env python3
"""Generate the tracker spreadsheet from data.json (the source of truth).

Usage: python3 build_xlsx_from_json.py [data.json] [out.xlsx] [tracker.html]

Produces three sheets:
  Tracker  — every application, one row each, section headers, live links
  Summary  — applications per day, where everything stands, how it was sent
  Glossary — the plain-English term list, lifted from tracker.html

Design law carried over from the site: colour is a redundant cue only.
Red and green are indistinguishable to a deuteranopic reader, so every
status cell carries its full text label and every status count is written
out in words. Nothing in this workbook is encoded by colour alone.
"""
import json
import re
import sys
from collections import Counter

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------- arguments
# CLI shape is unchanged: argv[1] = source json, argv[2] = output xlsx.
# argv[3] (new, optional) is the HTML file the glossary is lifted from.
# (pad per position — a single "argv + defaults" slice slides the defaults
#  along by one for every argument actually supplied)
_a = sys.argv[1:]
src      = _a[0] if len(_a) > 0 else 'data.json'
out      = _a[1] if len(_a) > 1 else 'Job_Application_Tracker.xlsx'
html_src = _a[2] if len(_a) > 2 else 'tracker.html'

# ------------------------------------------------------------------ palette
# Six semantic colours, matching the site. Each has a light tint that stays
# readable behind black text.
STATUS = {
    'rejected':  {'label': 'Rejected',
                  'meaning': 'They said no.',
                  'ink': 'F87171', 'tint': 'FBDDDD'},
    'submitted': {'label': 'Submitted, no reply',
                  'meaning': 'Sent. Nobody has acknowledged it.',
                  'ink': 'FBBF24', 'tint': 'FCEFCD'},
    'confirmed': {'label': 'Confirmed received',
                  'meaning': 'Their system said it arrived.',
                  'ink': '60A5FA', 'tint': 'DCEAFC'},
    'engaged':   {'label': 'A human engaged',
                  'meaning': 'Someone actually replied and talked.',
                  'ink': '34D399', 'tint': 'D2F4E7'},
    'bounced':   {'label': 'Never arrived',
                  'meaning': 'Bounced or withdrawn before it landed.',
                  'ink': '8B93A3', 'tint': 'E4E6EB'},
    'cold':      {'label': 'Cold outreach',
                  'meaning': 'No posted job; introduced myself anyway.',
                  'ink': 'C084FC', 'tint': 'EEDDFD'},
}
# Fixed display order for every status breakdown in the workbook.
STATUS_ORDER = ['rejected', 'submitted', 'confirmed', 'engaged', 'bounced', 'cold']

INK = '111827'        # body text — black-ish, for contrast on every tint
MUTED = '5B6472'
HEADER_BG = '1F2430'
RULE = 'D9DDE5'
GREY_TINT = STATUS['bounced']['tint']   # "Not recorded" is shown, in grey


def status_key(status):
    """Map a recorded status string to one of the six semantic keys.

    Order matters: an outcome always wins over the delivery receipt that
    preceded it. Never infer this from note text.
    """
    s = (status or '').lower()
    if 'bounced' in s or 'withdrawn' in s:
        return 'bounced'
    if 'rejected' in s or 'declined' in s:
        return 'rejected'
    if any(w in s for w in ('conversation', 'screen', 'interview', 'offer',
                            'loop', 'take-home', 'onsite')):
        return 'engaged'
    if 'cold' in s or 'emailed' in s:
        return 'cold'
    if '✓' in (status or ''):
        return 'confirmed'
    if 'submitted' in s:
        return 'submitted'
    return 'submitted'


# ------------------------------------------------------------------ channel
# How the application physically went out. Read from the note, first match
# wins. Notes often say "delivered; no bounce" — that is a delivery success,
# not a bounce, which is exactly why status is never inferred from notes.
CHANNEL_PATTERNS = [
    ('Direct email', r'direct email|cold email|emailed'),
    ('YC board', r'work at a startup|\bwaas\b|ycombinator|\byc\b'),
    ('Company application system',
     r'greenhouse|ashby|lever|workday|icims|polymer|rippling|wellfound|paraform'),
    # KNOWN GAP — see the note in the build report. Against today's data this
    # pattern never fires: all four notes containing "applied via" also name an
    # ATS (Polymer, Lever, Ashby, Workday), so the earlier rule claims them.
    # One row is plainly a web form — Rivet, "Google Form recorded" — and it
    # currently lands in "Not recorded". Adding |google form here yields the
    # 36/28/24/3/1 split the spec expects. NOT done unilaterally: this rule has
    # to stay identical to the one in tracker.html, so change both or neither.
    ('Web form',
     r'google form|applied via|application submitted|application form|web form|portal|careers page'),
]
CHANNEL_UNKNOWN = 'Not recorded'


def channel(note):
    n = (note or '').lower()
    for name, pattern in CHANNEL_PATTERNS:
        if re.search(pattern, n):
            return name
    return CHANNEL_UNKNOWN


# ----------------------------------------------------------------- glossary
def load_glossary(path):
    """Lift `const GLOSSARY = {...};` out of the site HTML.

    Fails loudly. An empty glossary sheet would look like a finished
    deliverable while silently being wrong, so we refuse to ship one.
    """
    try:
        html = open(path, encoding='utf-8').read()
    except OSError as e:
        raise SystemExit(
            f"GLOSSARY: cannot read '{path}' ({e}).\n"
            f"  The glossary is lifted from the site HTML. Pass its path as the\n"
            f"  third argument: build_xlsx_from_json.py data.json out.xlsx tracker.html"
        )
    m = re.search(r'const GLOSSARY = (\{.*?\});', html, re.S)
    if not m:
        raise SystemExit(
            f"GLOSSARY: no `const GLOSSARY = {{...}};` block found in '{path}'.\n"
            f"  Expected exactly that shape inside a <script> tag. Refusing to\n"
            f"  ship a workbook with an empty Glossary sheet."
        )
    try:
        data = json.loads(m.group(1))
    except json.JSONDecodeError as e:
        raise SystemExit(
            f"GLOSSARY: the GLOSSARY block in '{path}' is not valid JSON ({e}).\n"
            f"  Refusing to ship a workbook with an empty Glossary sheet."
        )
    if not isinstance(data, dict) or not data:
        raise SystemExit(
            f"GLOSSARY: the GLOSSARY block in '{path}' parsed to an empty object.\n"
            f"  Refusing to ship a workbook with an empty Glossary sheet."
        )
    return data


# --------------------------------------------------------------------- load
d = json.load(open(src, encoding='utf-8'))
meta = d['meta']
sb = meta['scoreboard']
glossary = load_glossary(html_src)

rows = [r for sec in d['sections'] for r in sec['rows']]
TOTAL = len(rows)

# --------------------------------------------------------------- style help
def F(**k):
    return Font(name='Arial', **k)


thin = Side(style='thin', color=RULE)
BOX = Border(bottom=thin)


def put(ws, r, c, value, *, bold=False, size=10, color=INK, fill=None,
        wrap=False, align='left', vert='top', italic=False, fmt=None):
    cell = ws.cell(r, c)
    cell.value = value
    cell.font = F(bold=bold, size=size, color=color, italic=italic)
    if fill:
        cell.fill = PatternFill('solid', fgColor=fill)
    cell.alignment = Alignment(wrap_text=wrap, vertical=vert, horizontal=align)
    if fmt:
        cell.number_format = fmt
    return cell


def header_row(ws, r, labels, start=1):
    for i, label in enumerate(labels):
        put(ws, r, start + i, label, bold=True, size=10, color='FFFFFF',
            fill=HEADER_BG, wrap=True)
    return r + 1


def set_widths(ws, widths, start=1):
    for i, w in enumerate(widths, start):
        ws.column_dimensions[get_column_letter(i)].width = w


wb = openpyxl.Workbook()

# ====================================================================== TRACKER
ws = wb.active
ws.title = 'Tracker'

COLS = ['Section', 'Company', 'Role', 'Salary', 'Location',
        'Status', 'As recorded', 'Plain meaning', 'How it was sent',
        'Date', 'Notes / Next']
NCOL = len(COLS)

r = header_row(ws, 1, COLS)
ws.freeze_panes = 'A2'

put(ws, r, 1, 'SCOREBOARD', bold=True)
put(ws, r, 2, f"{sb['applications_in']} applications in", bold=True)
put(ws, r, 3, f"{sb['ats_confirmations']} ATS confirmations (cumulative)", bold=True)
put(ws, r, 4, f"{sb['live_conversations']} live conversation", bold=True)
put(ws, r, 5, f"{sb['human_replies']} human reply · {sb['rejections']} rejections", bold=True)
put(ws, r, 6, f"{sb['in_pipeline']} in pipeline · {sb['backlog']} backlog", bold=True)
put(ws, r, 8, meta['counting_rule'], size=9, color=MUTED, wrap=True)
r += 1

put(ws, r, 1, 'HOW TO READ THIS', bold=True, size=9)
put(ws, r, 2, 'Status colour is a redundant cue only — every status cell is also '
              'written out in words, because red and green are indistinguishable '
              'to many readers. The full key, with counts, is on the Summary sheet.',
    size=9, color=MUTED, wrap=True)
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=NCOL)
ws.row_dimensions[r].height = 26
r += 2

for sec in d['sections']:
    put(ws, r, 1, sec['title'], bold=True, size=11, fill='F1F3F7')
    for c in range(2, NCOL + 1):
        put(ws, r, c, None, fill='F1F3F7')
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCOL)
    r += 1

    for row in sec['rows']:
        key = status_key(row['status'])
        meaning = STATUS[key]['meaning']
        tint = STATUS[key]['tint']
        values = [None, row['company'], row['role'], row['salary'], row['location'],
                  STATUS[key]['label'], row['status'], meaning,
                  channel(row['note']), row['date'], row['note']]
        for c, v in enumerate(values, 1):
            cell = put(ws, r, c, v, fill=tint, wrap=True)
            cell.border = BOX
        ws.cell(r, 6).font = F(size=10, bold=True, color=INK)
        if row.get('url'):
            link = ws.cell(r, 2)
            link.hyperlink = row['url']
            link.font = F(size=10, bold=True, color='1D4ED8', underline='single')
        r += 1

r += 1
put(ws, r, 1, 'STANDING RULES', bold=True)
put(ws, r, 2, d['standing_rules'], wrap=True)
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=NCOL)
r += 1
put(ws, r, 1, 'UPDATED', bold=True)
put(ws, r, 2, meta['updated'], wrap=True)
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=NCOL)

set_widths(ws, [16, 22, 30, 20, 22, 20, 14, 30, 22, 8, 52])

# ====================================================================== SUMMARY
sm = wb.create_sheet('Summary')
r = 1
put(sm, r, 1, meta['name'], bold=True, size=16)
r += 1
put(sm, r, 1, meta['subtitle'], size=10, color=MUTED, wrap=True)
sm.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
r += 1
put(sm, r, 1, meta['updated'], size=10, color=MUTED, wrap=True)
sm.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
sm.row_dimensions[r].height = 28
r += 2

# --- scoreboard
put(sm, r, 1, 'SCOREBOARD', bold=True, size=12)
r += 1
r = header_row(sm, r, ['Measure', 'Count'])
SCOREBOARD_LABELS = [
    ('applications_in', 'Applications in'),
    ('ats_confirmations', 'ATS confirmations (cumulative)'),
    ('live_conversations', 'Live conversations'),
    ('human_replies', 'Human replies'),
    ('rejections', 'Rejections'),
    ('in_pipeline', 'In pipeline'),
    ('backlog', 'Backlog'),
]
for k, label in SCOREBOARD_LABELS:
    put(sm, r, 1, label)
    put(sm, r, 2, sb[k], align='right')
    r += 1
r += 1

status_counts = Counter(status_key(row['status']) for row in rows)
confirmed_now = status_counts.get('confirmed', 0)

put(sm, r, 1, 'A NOTE ON THE ATS NUMBER', bold=True)
r += 1
put(sm, r, 1,
    f"The scoreboard says {sb['ats_confirmations']} ATS confirmations, but only "
    f"{confirmed_now} rows currently read “Submitted ✓”. Both are right. "
    f"{sb['ats_confirmations']} is cumulative — every confirmation ever received. "
    f"The other {sb['ats_confirmations'] - confirmed_now} (Rillet, Tsenta) were confirmed, "
    f"then moved on to a later outcome that now owns the row. The tables below are "
    f"current state, counted from the rows themselves.",
    size=10, color=MUTED, wrap=True)
sm.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
sm.row_dimensions[r].height = 46
r += 2

put(sm, r, 1, 'COUNTING RULE', bold=True)
r += 1
put(sm, r, 1, meta['counting_rule'], size=10, color=MUTED, wrap=True)
sm.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
sm.row_dimensions[r].height = 28
r += 3

# --- (a) applications per day
put(sm, r, 1, 'A. APPLICATIONS PER DAY', bold=True, size=12)
r += 1
put(sm, r, 1, 'Every dated row, oldest first. Dates are 2026.', size=9, color=MUTED)
r += 1
r = header_row(sm, r, ['Date', 'Date (long)', 'Applications', 'Running total'])

MONTHS = ['January', 'February', 'March', 'April', 'May', 'June', 'July',
          'August', 'September', 'October', 'November', 'December']


def date_sort_key(ds):
    try:
        mm, dd = ds.split('/')[:2]
        return (int(mm), int(dd))
    except (ValueError, IndexError):
        return (99, 99)


day_counts = Counter(row['date'] for row in rows)
running = 0
for ds in sorted(day_counts, key=date_sort_key):
    n = day_counts[ds]
    running += n
    mm, dd = date_sort_key(ds)
    long = f'{MONTHS[mm - 1]} {dd}, 2026' if mm <= 12 else ds
    put(sm, r, 1, ds)
    put(sm, r, 2, long, color=MUTED)
    put(sm, r, 3, n, align='right')
    put(sm, r, 4, running, align='right', color=MUTED)
    r += 1
put(sm, r, 1, 'Total', bold=True)
put(sm, r, 3, sum(day_counts.values()), bold=True, align='right')
r += 3

# --- (b) where they all stand
put(sm, r, 1, f'B. WHERE ALL {TOTAL} STAND TODAY', bold=True, size=12)
r += 1
put(sm, r, 1, 'This is also the colour key. Every status is named in words and '
              'counted — the colour adds nothing you cannot read.',
    size=9, color=MUTED, wrap=True)
sm.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
r += 1
r = header_row(sm, r, ['Status', 'Plain meaning', 'Count', '% of ' + str(TOTAL)])
for key in STATUS_ORDER:
    n = status_counts.get(key, 0)
    tint = STATUS[key]['tint']
    put(sm, r, 1, STATUS[key]['label'], bold=True, fill=tint)
    put(sm, r, 2, STATUS[key]['meaning'], fill=tint, wrap=True)
    put(sm, r, 3, n, align='right', fill=tint)
    put(sm, r, 4, (n / TOTAL) if TOTAL else 0, align='right', fill=tint, fmt='0.0%')
    r += 1
put(sm, r, 1, 'Total', bold=True)
put(sm, r, 3, sum(status_counts.values()), bold=True, align='right')
put(sm, r, 4, 1.0, bold=True, align='right', fmt='0.0%')
r += 3

# --- (c) how they were sent
put(sm, r, 1, 'C. HOW APPLICATIONS WERE SENT', bold=True, size=12)
r += 1
put(sm, r, 1, 'Read from each row’s note. “Not recorded” is shown, not '
              'hidden — an unknown channel is a real finding, not a gap to paper over.',
    size=9, color=MUTED, wrap=True)
sm.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
r += 1
r = header_row(sm, r, ['Channel', 'Count', '% of ' + str(TOTAL)])

channel_counts = Counter(channel(row['note']) for row in rows)
known = sorted(((c, n) for c, n in channel_counts.items() if c != CHANNEL_UNKNOWN),
               key=lambda kv: (-kv[1], kv[0]))
ordered = known + ([(CHANNEL_UNKNOWN, channel_counts[CHANNEL_UNKNOWN])]
                   if CHANNEL_UNKNOWN in channel_counts else [])
for name, n in ordered:
    fill = GREY_TINT if name == CHANNEL_UNKNOWN else None
    put(sm, r, 1, name, fill=fill, bold=(name == CHANNEL_UNKNOWN))
    put(sm, r, 2, n, align='right', fill=fill)
    put(sm, r, 3, (n / TOTAL) if TOTAL else 0, align='right', fill=fill, fmt='0.0%')
    r += 1
put(sm, r, 1, 'Total', bold=True)
put(sm, r, 2, sum(channel_counts.values()), bold=True, align='right')
put(sm, r, 3, 1.0, bold=True, align='right', fmt='0.0%')
r += 3

put(sm, r, 1, 'STANDING RULES', bold=True)
r += 1
put(sm, r, 1, d['standing_rules'], size=10, color=MUTED, wrap=True)
sm.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
sm.row_dimensions[r].height = 28

set_widths(sm, [34, 44, 14, 14, 14])

# ===================================================================== GLOSSARY
gl = wb.create_sheet('Glossary')
r = 1
put(gl, r, 1, 'Glossary', bold=True, size=16)
r += 1
put(gl, r, 1, 'Plain-English definitions, lifted from the site so the two can '
              'never drift apart.', size=10, color=MUTED, wrap=True)
gl.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
r += 2
r = header_row(gl, r, ['Group', 'Term', 'Definition'])

for group, entries in glossary.items():
    put(gl, r, 1, group, bold=True, size=11, fill='F1F3F7')
    put(gl, r, 2, None, fill='F1F3F7')
    put(gl, r, 3, None, fill='F1F3F7')
    r += 1
    for entry in entries:
        term = entry.get('term', '')
        aliases = entry.get('aliases') or []
        if aliases:
            term = f"{term} (also: {', '.join(aliases)})"
        put(gl, r, 1, None)
        put(gl, r, 2, term, bold=True, wrap=True).border = BOX
        put(gl, r, 3, entry.get('definition', ''), wrap=True).border = BOX
        r += 1

gl.freeze_panes = 'A5'
set_widths(gl, [26, 30, 88])

wb.save(out)
print(f'saved {out} — {TOTAL} applications, '
      f'{len(glossary)} glossary groups, 3 sheets')
